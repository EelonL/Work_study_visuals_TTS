"""
Työajan havainnointi – Kuvaajageneraattori
========================================

Lukee yhden tai useamman Excel-tiedoston "Havainnot"-välilehdeltä
ja piirtää jokaiselle henkilölle omat kuvaajat:

  1. Aikajanakaavio  – Tekemisaika yläpuolella, Apuaika alapuolella,
                       Häiriöaika ja muu aika omilla väreillään.
  2. Työneräkaavio   – Tekemisaika yläpuolella, muut ajat alapuolella.
                       Työnerät esitetään omilla väreillään palkkeina.
                       Lisäksi näytetään työneräkohtaiset jaksojen pituustilastot.
  3. Yhteenvetokuvaaja – Aikalajien prosenttiosuudet päivittäin.

Käyttö:
  Streamlit:   streamlit run tyoajan_havainnointi_kuvaajat.py
  Paikallinen: python tyoajan_havainnointi_kuvaajat.py

Vaatimukset (requirements.txt):
  openpyxl
  matplotlib
  streamlit
  pandas
"""

import math
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl
import pandas as pd

# ── Tunnistetaan ajoympäristö ──────────────────────────────────────────────

try:
    import streamlit as st
    STREAMLIT = True
except ImportError:
    STREAMLIT = False

# ── Asetukset ──────────────────────────────────────────────────────────────

GAP_BETWEEN_DAYS = 20

MARKER_SIZE_HAIRIO = 600
MARKER_SIZE_MUU = 500

COLORS = {
    "Tekemisaika": "#1976D2",
    "Apuaika":     "#FB8C00",
    "Häiriöaika":  "#E53935",
    "Muu":         "#8E24AA",
    "Valmiusaika": "#43A047",
    "Taukoaika":   "#757575",
    "Tuntematon":  "#9E9E9E",
}

SUMMARY_CATS = [
    "Tekemisaika", "Apuaika", "Häiriöaika",
    "Valmiusaika", "Taukoaika", "Muu", "Tuntematon",
]

PERSON_COLUMNS = {
    2: "Henkilö 1",  # C
    3: "Henkilö 2",  # D
    4: "Henkilö 3",  # E
    5: "Henkilö 4",  # F
}

# ── Luokittelufunktio ──────────────────────────────────────────────────────

def classify_code(code) -> str:
    """Palauttaa aikalajin havaintokoodin perusteella."""
    try:
        c = float(str(code).replace(",", "."))
    except (ValueError, TypeError):
        return "Tuntematon"

    if 1 <= c < 40:
        return "Tekemisaika"
    if 40 <= c < 50:
        return "Apuaika"
    if c == 50:
        return "Valmiusaika"
    if 51 <= c <= 55:
        return "Häiriöaika"
    if 56 <= c <= 59:
        return "Muu"
    if c == 60:
        return "Taukoaika"
    return "Tuntematon"


STOPWORDS_FI = {
    "ja", "tai", "sekä", "myös", "samalla", "kanssa", "että", "oli", "lopulta",
    "käyttää", "käyttääkö", "käytti", "katsoo", "katsoo", "lukee", "puhuu",
    "soittaa", "kirjoittaa", "täyttää", "edelleen", "jolla", "jossa", "jonka",
    "the", "of", "mm", "ei", "on", "oli", "kun", "jos", "esim", "sis",
    "työ", "työasia", "työpisteessä"
}


def normalize_text(value) -> str:
    text = str(value or "").lower()
    text = text.replace("å", "a").replace("ä", "a").replace("ö", "o")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def text_tokens(value) -> list:
    tokens = []
    for tok in normalize_text(value).split():
        if len(tok) < 3:
            continue
        if tok in STOPWORDS_FI:
            continue
        tokens.append(tok)
    return tokens


def token_roots(tokens: list) -> set:
    roots = set()
    for tok in tokens:
        roots.add(tok)
        if len(tok) >= 6:
            roots.add(tok[:5])
        elif len(tok) >= 4:
            roots.add(tok[:4])
    return roots


def build_catalog_match_index(batch_catalog: dict) -> dict:
    index = {}
    for code_key, info in batch_catalog.items():
        name = str(info.get("name", "")).strip()
        if not name:
            continue
        tokens = text_tokens(name)
        index[code_key] = {
            "tokens": tokens,
            "roots": token_roots(tokens),
            "name_norm": normalize_text(name),
            "category": str(info.get("category", "")).strip(),
            "name": name,
        }
    return index


def score_note_against_item(note: str, item: dict) -> float:
    note_norm = normalize_text(note)
    note_tokens = text_tokens(note)
    note_roots = token_roots(note_tokens)

    score = 0.0
    overlap = note_roots.intersection(item["roots"])
    score += 2.5 * len(overlap)

    for tok in item["tokens"]:
        if len(tok) >= 4 and tok in note_norm:
            score += 1.5

    # Hyvin lyhyet mutta tärkeät osumat
    if "excel" in note_norm and "excel" in item["name_norm"]:
        score += 2
    if "teams" in note_norm and "teams" in item["name_norm"]:
        score += 2
    if "winbus" in note_norm and "winbus" in item["name_norm"]:
        score += 2
    if "whatsapp" in note_norm and "whatsapp" in item["name_norm"]:
        score += 2
    if "puhelu" in note_norm and "puhelu" in item["name_norm"]:
        score += 2
    if "sahkopost" in note_norm and "sahkopost" in item["name_norm"]:
        score += 2
    if "kuljettaj" in note_norm and "kuljettaj" in item["name_norm"]:
        score += 2
    if "keskustel" in note_norm and "keskustel" in item["name_norm"]:
        score += 1

    return score


def find_digit_segmentations(code_key, valid_codes: set, max_parts: int = 4) -> list:
    code_str = str(code_key).replace(".0", "")
    if not code_str.isdigit():
        return []

    results = []

    def rec(pos: int, parts: list):
        if pos == len(code_str):
            if 2 <= len(parts) <= max_parts:
                results.append(parts.copy())
            return
        if len(parts) >= max_parts:
            return

        for length in (2, 1):
            part = code_str[pos:pos + length]
            if not part:
                continue
            if part.startswith("0"):
                continue
            part_int = int(part)
            if part_int in valid_codes:
                parts.append(part_int)
                rec(pos + length, parts)
                parts.pop()

    rec(0, [])
    unique = []
    seen = set()
    for seg in results:
        tup = tuple(seg)
        if tup not in seen:
            seen.add(tup)
            unique.append(seg)
    return unique


def choose_best_segmentation(code_key, note: str, batch_catalog: dict, catalog_index: dict):
    valid_codes = {k for k in batch_catalog.keys() if isinstance(k, int) and 1 <= k <= 60}
    segmentations = find_digit_segmentations(code_key, valid_codes)
    if not segmentations:
        return []

    best = []
    best_score = -1.0
    for seg in segmentations:
        score = 0.0
        for part in seg:
            item = catalog_index.get(part)
            if item:
                score += score_note_against_item(note, item)
        # Pieni bonus kahden tai kolmen osan selityksille
        if 2 <= len(seg) <= 3:
            score += 0.5
        if score > best_score:
            best_score = score
            best = seg

    return best if best_score > 0 else []


def infer_combined_code_info(code_key, notes: list, batch_catalog: dict, catalog_index: dict) -> dict:
    """
    Päättelee yhdistelmäkoodille nimen ja aikalajin.
    Tavoite:
    - käyttää ensisijaisesti Eräluettelo-välilehden nimiä
    - jos se ei onnistu, käyttää huomiosarakkeen yleisintä tekstiä
    - ei luokittele tuntematonta yhdistelmäkoodia tauoksi
    """
    notes = [str(n).strip() for n in notes if str(n or "").strip()]
    note_counter = Counter(notes)
    main_note = note_counter.most_common(1)[0][0] if note_counter else ""

    seg = choose_best_segmentation(code_key, main_note, batch_catalog, catalog_index)
    if seg:
        names = []
        categories = []
        for part in seg:
            info = batch_catalog.get(part, {})
            name = str(info.get("name", "")).strip()
            if name:
                names.append(name)
            cat = str(info.get("category", "")).strip()
            if cat:
                categories.append(cat)

        if names:
            category = categories[0] if categories and len(set(categories)) == 1 else (categories[0] if categories else "Tekemisaika")
            return {
                "name": " + ".join(names),
                "category": category or "Tekemisaika",
                "synthetic": True,
                "source": "digit_segmentation"
            }

    # Tekstihaku Eräluettelon nimistä
    scored = []
    for base_code, item in catalog_index.items():
        sc = score_note_against_item(main_note, item)
        if sc >= 3.5:
            scored.append((sc, base_code))
    scored.sort(reverse=True)

    chosen = []
    used = set()
    for sc, base_code in scored:
        if base_code in used:
            continue
        chosen.append((sc, base_code))
        used.add(base_code)
        if len(chosen) >= 2:
            break

    note_looks_combined = any(marker in normalize_text(main_note) for marker in [" samalla ", " ja ", "+"])

    if chosen:
        top_score = chosen[0][0]
        chosen_codes = [base_code for _, base_code in chosen]

        # Jos huomio-teksti selvästi kuvaa yhdistelmätekemistä, mutta löytyi vain
        # yksi heikko osuma, käytetään mieluummin huomio-tekstiä sellaisenaan.
        if not (len(chosen_codes) == 1 and note_looks_combined and top_score < 6.0):
            names = []
            categories = []
            for base_code in chosen_codes:
                info = batch_catalog.get(base_code, {})
                nm = str(info.get("name", "")).strip()
                if nm:
                    names.append(nm)
                cat = str(info.get("category", "")).strip()
                if cat:
                    categories.append(cat)

            allow_note_match = (
                len(chosen_codes) >= 2 or
                (len(chosen_codes) == 1 and (not note_looks_combined) and top_score >= 7.5)
            )
            if names and allow_note_match and top_score >= 4.5:
                category = categories[0] if categories and len(set(categories)) == 1 else (categories[0] if categories else "Tekemisaika")
                return {
                    "name": " + ".join(names),
                    "category": category or "Tekemisaika",
                    "synthetic": True,
                    "source": "note_match"
                }

    # Viimeinen turvallinen fallback: käytetään huomio-tekstiä työnerän nimenä
    if main_note:
        return {
            "name": main_note,
            "category": "Tekemisaika",
            "synthetic": True,
            "source": "note_fallback"
        }

    return {
        "name": "",
        "category": "Tuntematon",
        "synthetic": True,
        "source": "unknown"
    }


# ── Datan luku ─────────────────────────────────────────────────────────────

def read_batch_catalog(wb) -> dict:
    """
    Lukee välilehden "Eräluettelo" ja palauttaa sanakirjan:
      {code: {"name": "...", "category": "..."}}

    Mukaan otetaan vain ne erät, joilla on erän nimi.
    """
    if "Eräluettelo" not in wb.sheetnames:
        return {}

    ws = wb["Eräluettelo"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header_idx = None
    col_code = None
    col_name = None
    col_cat = None

    for i, row in enumerate(rows):
        if not row:
            continue
        normalized = [str(v).strip() if v is not None else "" for v in row]
        if "Eränro" in normalized and "Erän nimi" in normalized:
            header_idx = i
            col_code = normalized.index("Eränro")
            col_name = normalized.index("Erän nimi")
            col_cat = normalized.index("Aikalaji") if "Aikalaji" in normalized else None
            break

    if header_idx is None:
        return {}

    batch_catalog = {}
    for row in rows[header_idx + 1:]:
        if not row:
            continue

        code = row[col_code] if col_code < len(row) else None
        name = row[col_name] if col_name < len(row) else None
        cat = row[col_cat] if (col_cat is not None and col_cat < len(row)) else None

        if code is None or name is None:
            continue

        name_str = str(name).strip()
        if not name_str:
            continue

        try:
            code_num = float(str(code).replace(",", "."))
            code_key = int(code_num) if code_num.is_integer() else code_num
        except Exception:
            code_key = str(code).strip()

        batch_catalog[code_key] = {
            "name": name_str,
            "category": str(cat).strip() if cat is not None else "",
        }

    return batch_catalog


def normalize_code_key(code):
    try:
        code_num = float(str(code).replace(",", "."))
        return int(code_num) if code_num.is_integer() else code_num
    except Exception:
        return str(code).strip()


def read_file(filepath) -> dict:
    """
    Lukee yhden Excel-tiedoston havaintodatan.

    Palauttaa muodon:
    {
        "date": <date|None>,
        "persons": {
            "Henkilö 1": [obs, obs, ...],
            ...
        },
        "batch_catalog": {code: {"name": ..., "category": ...}}
    }
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        raise IOError(f"Tiedoston avaaminen epäonnistui: {e}")

    if "Havainnot" not in wb.sheetnames:
        wb.close()
        raise ValueError("Tiedostosta puuttuu välilehti 'Havainnot'.")

    batch_catalog = read_batch_catalog(wb)

    ws = wb["Havainnot"]
    rows = list(ws.iter_rows(values_only=True))

    date_val = rows[2][5] if len(rows) > 2 and len(rows[2]) > 5 else None
    meas_date = date_val.date() if isinstance(date_val, datetime) else None

    if len(rows) < 6:
        wb.close()
        return {"date": meas_date, "persons": {}, "batch_catalog": batch_catalog}

    person_observations = defaultdict(list)

    for row in rows[5:]:
        if len(row) < 2 or row[0] is None or row[1] is None:
            continue

        try:
            h = int(row[0])
            m = int(row[1])
        except (TypeError, ValueError):
            continue

        for col_idx, person_name in PERSON_COLUMNS.items():
            code = row[col_idx] if col_idx < len(row) else None
            if code is None or str(code).strip() == "":
                continue

            note = row[12] if len(row) > 12 else None
            person_observations[person_name].append({
                "hour": h,
                "minute": m,
                "code": code,
                "code_key": normalize_code_key(code),
                "note": str(note).strip() if note is not None else "",
                "category": classify_code(code),
            })

    wb.close()
    return {
        "date": meas_date,
        "persons": dict(person_observations),
        "batch_catalog": batch_catalog,
    }


# ── Apufunktiot ────────────────────────────────────────────────────────────

def build_segments(series: list, category: str) -> list:
    segs = []
    in_seg = False
    seg_start = None
    prev_x = None

    for obs in series:
        if obs["category"] == category:
            if not in_seg:
                seg_start = obs["x"]
                in_seg = True
            prev_x = obs["x"]
        else:
            if in_seg:
                segs.append((seg_start, prev_x + 1))
                in_seg = False

    if in_seg and prev_x is not None:
        segs.append((seg_start, prev_x + 1))

    return segs


def minutes_to_label(abs_minutes: float) -> str:
    h = int(abs_minutes // 60) % 24
    m = int(abs_minutes % 60)
    return f"{h:02d}:{m:02d}"


def build_day_info(datasets: list) -> list:
    day_info = []
    x_offset = 0

    for ds in datasets:
        obs = ds["observations"]
        if not obs:
            continue

        obs = sorted(obs, key=lambda o: (o["hour"], o["minute"]))
        start_abs = obs[0]["hour"] * 60 + obs[0]["minute"]
        series = []

        for o in obs:
            abs_min = o["hour"] * 60 + o["minute"]
            x = x_offset + (abs_min - start_abs)
            series.append({**o, "x": x, "abs_min": abs_min})

        end_x = series[-1]["x"]
        day_info.append({
            "date": ds["date"],
            "x_start": x_offset,
            "x_end": end_x + 1,
            "x_mid": (x_offset + end_x) / 2,
            "start_abs": start_abs,
            "series": series,
        })
        x_offset = end_x + 1 + GAP_BETWEEN_DAYS

    return day_info


def collect_combined_code_candidates(file_datasets: list):
    """
    Kerää yhdistelmäkoodiehdokkaat ja ohjelman alustavat arviot ennen laskentaa.
    Palauttaa:
      - merged_batch_catalog: Eräluetteloista yhdistetty perusluettelo
      - candidates: lista ehdokasrivejä käyttöliittymää varten
      - auto_catalog: ohjelman automaattiset tulkinnat yhdistelmäkoodeille
    """
    merged_batch_catalog = {}

    for ds in file_datasets:
        for code_key, info in ds.get("batch_catalog", {}).items():
            if code_key not in merged_batch_catalog:
                merged_batch_catalog[code_key] = info

    catalog_index = build_catalog_match_index(merged_batch_catalog)

    unknown_notes = defaultdict(list)
    for ds in file_datasets:
        for observations in ds["persons"].values():
            for obs in observations:
                code_key = obs["code_key"]
                if code_key not in merged_batch_catalog:
                    unknown_notes[code_key].append(obs.get("note", ""))

    candidates = []
    auto_catalog = {}
    for code_key in sorted(unknown_notes.keys(), key=sort_code_key):
        notes = [str(n).strip() for n in unknown_notes[code_key] if str(n or "").strip()]
        inferred = infer_combined_code_info(code_key, notes, merged_batch_catalog, catalog_index)
        auto_catalog[code_key] = inferred

        note_counter = Counter(notes)
        example_note = note_counter.most_common(1)[0][0] if note_counter else ""
        source = inferred.get("source", "")
        source_label_map = {
            "digit_segmentation": "Numeropilkonta + Eräluettelo",
            "note_match": "Huomio-teksti + Eräluettelo",
            "note_fallback": "Huomio-teksti",
            "unknown": "Ei varmaa tulkintaa",
        }

        candidates.append({
            "code_key": code_key,
            "code_text": str(code_key).replace(".0", ""),
            "suggested_name": str(inferred.get("name", "")).strip(),
            "suggested_category": str(inferred.get("category", "Tuntematon")).strip() or "Tuntematon",
            "source": source,
            "source_label": source_label_map.get(source, source or "Arvio"),
            "example_note": example_note,
            "note_count": len(notes),
        })

    return merged_batch_catalog, candidates, auto_catalog


def build_person_day_infos(file_datasets: list, catalog_overrides: dict | None = None):
    """
    Henkilöt yhdistetään tiedostojen välillä sarakepaikan perusteella:
      C = Henkilö 1, D = Henkilö 2, E = Henkilö 3, F = Henkilö 4

    Lisäksi yhdistelmäkoodeille käytetään joko:
    - käyttäjän vahvistamia tulkintoja, tai
    - ohjelman automaattisia arvioita.
    """
    person_days = defaultdict(list)

    merged_batch_catalog, _, auto_catalog = collect_combined_code_candidates(file_datasets)
    effective_catalog = dict(merged_batch_catalog)
    effective_catalog.update(auto_catalog)

    if catalog_overrides:
        for code_key, info in catalog_overrides.items():
            effective_catalog[code_key] = {
                **effective_catalog.get(code_key, {}),
                **info,
            }

    for ds in file_datasets:
        date_val = ds["date"]
        for person_name, observations in ds["persons"].items():
            if not observations:
                continue

            resolved_obs = []
            for obs in observations:
                code_key = obs["code_key"]
                info = effective_catalog.get(code_key, {})
                resolved_category = str(info.get("category", "")).strip() or obs.get("category", "Tuntematon")
                resolved_obs.append({
                    **obs,
                    "category": resolved_category,
                })

            person_days[person_name].append({
                "date": date_val,
                "observations": resolved_obs,
            })

    for person_name in person_days:
        person_days[person_name].sort(key=lambda d: d["date"] or datetime.min.date())

    person_day_infos = {
        person_name: build_day_info(day_datasets)
        for person_name, day_datasets in person_days.items()
    }
    return person_day_infos, effective_catalog


def get_xticks_for_day_info(day_info: list):
    tick_positions = []
    tick_labels = []
    end_tick_positions = set()

    for day in day_info:
        duration = day["x_end"] - day["x_start"]
        for offset in range(0, int(duration) + 1, 60):
            tick_positions.append(day["x_start"] + offset)
            tick_labels.append(minutes_to_label(day["start_abs"] + offset))

        last_series = day["series"][-1]
        end_abs = last_series["abs_min"]
        end_x = last_series["x"]

        if end_x not in tick_positions:
            tick_positions.append(end_x)
            tick_labels.append(minutes_to_label(end_abs))
            end_tick_positions.add(end_x)

    ticks_sorted = sorted(zip(tick_positions, tick_labels), key=lambda t: t[0])
    if not ticks_sorted:
        return [], [], set()

    tick_positions = [t[0] for t in ticks_sorted]
    tick_labels = [t[1] for t in ticks_sorted]
    return tick_positions, tick_labels, end_tick_positions


def sort_code_key(code):
    try:
        return float(str(code).replace(",", "."))
    except Exception:
        return str(code)


def format_batch_label(code_key, batch_catalog):
    info = batch_catalog.get(code_key, {})
    name = str(info.get("name", "")).strip()
    code_text = str(code_key).replace(".0", "")
    return f"{code_text} – {name}" if name else f"Työnerä {code_text}"


def build_chart3_legend_df(day_info: list, batch_catalog: dict) -> pd.DataFrame:
    """
    Rakentaa työneräkaavion selitteen datamuotoon.
    """
    used_codes = set()

    for day in day_info:
        for run in get_code_runs(day["series"]):
            if run["category"] in ("Tekemisaika", "Apuaika", "Valmiusaika") and run["code_key"] in batch_catalog:
                used_codes.add(run["code_key"])

    rows = []
    for code_key in sorted(used_codes, key=sort_code_key):
        info = batch_catalog.get(code_key, {})
        rows.append({
            "Koodi": str(code_key).replace(".0", ""),
            "Selite": str(info.get("name", "")).strip() or format_batch_label(code_key, batch_catalog),
            "Ryhmä": str(info.get("category", "")).strip() or "Työnerä",
        })

    rows.extend([
        {"Koodi": "-", "Selite": "Häiriöaika", "Ryhmä": "Kiinteä väri"},
        {"Koodi": "-", "Selite": "Muu", "Ryhmä": "Kiinteä väri"},
        {"Koodi": "-", "Selite": "Taukoaika", "Ryhmä": "Kiinteä väri"},
        {"Koodi": "-", "Selite": "Tuntematon / nimeämätön erä", "Ryhmä": "Kiinteä väri"},
        {"Koodi": "-", "Selite": "Mittauksen aloitus", "Ryhmä": "Merkintä"},
        {"Koodi": "-", "Selite": "Mittauksen päättyminen", "Ryhmä": "Merkintä"},
    ])

    return pd.DataFrame(rows)


def make_chart3_legend_figure(day_info: list, batch_catalog: dict):
    """
    Piirtää työneräkaavion selitteen omana erillisenä kuvanaan,
    jotta värikytkentä säilyy mutta varsinainen kuvaaja pysyy puhtaana.
    """
    used_codes = set()
    for day in day_info:
        for run in get_code_runs(day["series"]):
            if run["category"] in ("Tekemisaika", "Apuaika", "Valmiusaika") and run["code_key"] in batch_catalog:
                used_codes.add(run["code_key"])

    unique_codes = sorted(used_codes, key=sort_code_key)
    cmap = plt.get_cmap("tab20")
    code_colors = {code_key: cmap(i % 20) for i, code_key in enumerate(unique_codes)}

    legend_items = []
    for code_key in unique_codes:
        legend_items.append(("patch", code_colors[code_key], format_batch_label(code_key, batch_catalog)))

    legend_items.extend([
        ("patch", COLORS["Häiriöaika"], "Häiriöaika"),
        ("patch", COLORS["Muu"], "Muu"),
        ("patch", COLORS["Taukoaika"], "Taukoaika"),
        ("patch", COLORS["Tuntematon"], "Tuntematon / nimeämätön erä"),
        ("line_start", "#888888", "Mittauksen aloitus"),
        ("line_end", "#C62828", "Mittauksen päättyminen"),
    ])

    n_items = max(1, len(legend_items))
    ncols = 2 if n_items <= 12 else 3
    nrows = (n_items + ncols - 1) // ncols

    fig_h = max(1.6, 0.52 * nrows + 0.55)
    fig, ax = plt.subplots(figsize=(22, fig_h))
    fig.patch.set_facecolor("#F5F5F5")
    ax.set_facecolor("#FAFAFA")
    ax.set_xlim(0, ncols)
    ax.set_ylim(0, nrows)
    ax.axis("off")

    for i, (kind, color, label) in enumerate(legend_items):
        col = i // nrows
        row = i % nrows
        y = nrows - row - 0.55
        x = col + 0.06

        if kind == "patch":
            rect = mpatches.Rectangle((x, y - 0.10), 0.10, 0.20, facecolor=color, edgecolor="#FFFFFF", linewidth=0.8)
            ax.add_patch(rect)
        elif kind == "line_start":
            ax.plot([x, x + 0.12], [y, y], color=color, linestyle="--", linewidth=2.0, solid_capstyle="butt")
        elif kind == "line_end":
            ax.plot([x, x + 0.12], [y, y], color=color, linestyle=":", linewidth=2.2, solid_capstyle="butt")

        ax.text(x + 0.15, y, label, va="center", ha="left", fontsize=8.5, color="#222222")

    fig.subplots_adjust(left=0.02, right=0.98, top=0.96, bottom=0.08)
    return fig

def get_code_runs(series):
    if not series:
        return []

    runs = []
    current = series[0]
    run_start_x = current["x"]
    prev_x = current["x"]
    run_len = 1

    for obs in series[1:]:
        if obs["code_key"] == current["code_key"]:
            prev_x = obs["x"]
            run_len += 1
        else:
            runs.append({
                "x0": run_start_x,
                "x1": prev_x + 1,
                "length": run_len,
                "category": current["category"],
                "code": current["code"],
                "code_key": current["code_key"],
            })
            current = obs
            run_start_x = obs["x"]
            prev_x = obs["x"]
            run_len = 1

    runs.append({
        "x0": run_start_x,
        "x1": prev_x + 1,
        "length": run_len,
        "category": current["category"],
        "code": current["code"],
        "code_key": current["code_key"],
    })
    return runs


def compute_batch_run_statistics(day_info: list, batch_catalog: dict) -> pd.DataFrame:
    """
    Laskee työneräkohtaiset yhtäjaksoisten jaksojen pituustilastot.

    Mukaan otetaan vain:
    - sellaiset erät, joilla on nimi Eräluettelo-välilehdellä
    - tekemisaikaan kuuluvat jaksot
    - taulukko järjestetään keskiarvon mukaan laskevaan järjestykseen
    """
    lengths_by_code = defaultdict(list)

    for day in day_info:
        for run in get_code_runs(day["series"]):
            code_key = run["code_key"]
            if code_key not in batch_catalog:
                continue
            if run["category"] != "Tekemisaika":
                continue
            lengths_by_code[code_key].append(run["length"])

    rows = []
    for code_key in lengths_by_code.keys():
        lengths = lengths_by_code[code_key]
        if not lengths:
            continue

        info = batch_catalog.get(code_key, {})
        mean_val = sum(lengths) / len(lengths)
        rows.append({
            "Eränro": str(code_key).replace(".0", ""),
            "Erän nimi": info.get("name", ""),
            "Jaksoja (n)": len(lengths),
            "Min (min)": int(min(lengths)),
            "Max (min)": int(max(lengths)),
            "Keskiarvo (min)": round(mean_val, 2),
            "Keskihajonta": round(pd.Series(lengths).std(ddof=1), 2) if len(lengths) > 1 else 0.0,
        })

    if not rows:
        return pd.DataFrame(columns=[
            "Eränro", "Erän nimi", "Jaksoja (n)", "Min (min)",
            "Max (min)", "Keskiarvo (min)", "Keskihajonta"
        ])

    df = pd.DataFrame(rows)
    df = df.sort_values(by=["Keskiarvo (min)", "Max (min)", "Eränro"], ascending=[False, False, True])
    df = df.reset_index(drop=True)
    return df


# ── Kuvaajien piirto ───────────────────────────────────────────────────────

def make_chart1(day_info: list, person_name: str = ""):
    fig, ax = plt.subplots(figsize=(20, 4))
    fig.patch.set_facecolor("#F5F5F5")
    ax.set_facecolor("#FAFAFA")

    first_day = True
    for day in day_info:
        series = day["series"]

        for x0, x1 in build_segments(series, "Tekemisaika"):
            ax.broken_barh([(x0, x1 - x0)], (0, 1), facecolor=COLORS["Tekemisaika"], alpha=0.85,
                           label="Tekemisaika" if first_day else "")

        for x0, x1 in build_segments(series, "Apuaika"):
            ax.broken_barh([(x0, x1 - x0)], (-1, 1), facecolor=COLORS["Apuaika"], alpha=0.85,
                           label="Apuaika" if first_day else "")

        for x0, x1 in build_segments(series, "Valmiusaika"):
            ax.broken_barh([(x0, x1 - x0)], (0, 1), facecolor=COLORS["Valmiusaika"], alpha=0.85,
                           label="Valmiusaika" if first_day else "")

        for x0, x1 in build_segments(series, "Taukoaika"):
            ax.broken_barh([(x0, x1 - x0)], (-1, 1), facecolor=COLORS["Taukoaika"], alpha=0.70,
                           label="Taukoaika" if first_day else "")

        hairio = [o["x"] + 0.5 for o in series if o["category"] == "Häiriöaika"]
        if hairio:
            ax.scatter(hairio, [0] * len(hairio), marker=(10, 1, 0), s=MARKER_SIZE_HAIRIO,
                       color=COLORS["Häiriöaika"], zorder=6, label="Häiriöaika" if first_day else "",
                       linewidths=0.5, edgecolors="white")

        muu = [o["x"] + 0.5 for o in series if o["category"] == "Muu"]
        if muu:
            ax.scatter(muu, [0] * len(muu), marker="o", s=MARKER_SIZE_MUU,
                       color=COLORS["Muu"], zorder=6, label="Muu" if first_day else "",
                       linewidths=0.8, edgecolors="white")

        first_day = False

    tick_positions, tick_labels, end_tick_positions = get_xticks_for_day_info(day_info)
    ax.set_xticks(tick_positions)
    ax.set_xticklabels(tick_labels, rotation=45, ha="right", fontsize=8)

    for label, pos in zip(ax.get_xticklabels(), tick_positions):
        if pos in end_tick_positions:
            label.set_color("#C62828")
            label.set_fontweight("bold")
        else:
            label.set_color("#333333")
            label.set_fontweight("normal")

    for day in day_info:
        end_x = day["series"][-1]["x"]
        ax.axvline(end_x, color="#C62828", linestyle=":", linewidth=1.2, zorder=4)

    for day in day_info:
        date_label = day["date"].strftime("%d.%m.%Y") if day["date"] else "?"
        x_start = day["x_start"]
        ax.axvline(x_start, color="#888888", linestyle="--", linewidth=1.2, zorder=4)
        ax.text(x_start + 1.0, -1.45, date_label, rotation=90, rotation_mode="anchor",
                ha="left", va="bottom", fontsize=9, fontweight="bold", color="#444444", zorder=5)

    if len(day_info) > 1:
        for i in range(len(day_info) - 1):
            gap_x = (day_info[i]["x_end"] + day_info[i + 1]["x_start"]) / 2
            ax.axvline(gap_x, color="#BBBBBB", linestyle="--", linewidth=1.0, zorder=2)

    ax.axhline(0, color="#333333", linewidth=1.2, zorder=3)
    ax.set_ylim(-1.6, 2.0)
    ax.set_yticks([-0.5, 0.5])
    ax.set_yticklabels(["Apuaika", "Tekemisaika"], fontsize=9)
    ax.tick_params(axis="y", length=0)

    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.grid(axis="x", color="#E0E0E0", linestyle="-", linewidth=0.5, zorder=0)

    legend_handles = [
        mpatches.Patch(color=COLORS["Tekemisaika"], label="Tekemisaika"),
        mpatches.Patch(color=COLORS["Apuaika"], label="Apuaika"),
        mpatches.Patch(color=COLORS["Valmiusaika"], label="Valmiusaika"),
        mpatches.Patch(color=COLORS["Taukoaika"], label="Taukoaika"),
        plt.Line2D([0], [0], marker=(10, 1, 0), color="w", markerfacecolor=COLORS["Häiriöaika"],
                   markersize=14, label="Häiriöaika"),
        plt.Line2D([0], [0], marker="o", color="w", markerfacecolor=COLORS["Muu"],
                   markersize=12, label="Muu"),
        plt.Line2D([0], [0], color="#888888", linestyle="--", linewidth=1.5, label="Mittauksen aloitus"),
        plt.Line2D([0], [0], color="#C62828", linestyle=":", linewidth=1.5, label="Mittauksen päättyminen"),
    ]
    ax.legend(handles=legend_handles, loc="lower right", framealpha=0.9, fontsize=8, ncol=4)

    title = "Aikalajit ajan suhteen – minuuttihavainnot"
    if person_name:
        title += f" ({person_name})"
    ax.set_title(title, fontsize=13, fontweight="bold", pad=30)
    ax.set_xlabel("Kellonaika", fontsize=9)
    plt.tight_layout()
    return fig


def make_chart2(day_info: list, person_name: str = ""):
    n_days = len(day_info)
    fig, axes = plt.subplots(1, n_days, figsize=(max(5 * n_days, 6), 5), sharey=False, squeeze=False)
    fig.patch.set_facecolor("#F5F5F5")

    title = "Aikalajien osuudet päivittäin"
    if person_name:
        title += f" ({person_name})"
    fig.suptitle(title, fontsize=13, fontweight="bold", y=1.02)

    for col_i, day in enumerate(day_info):
        ax = axes[0][col_i]
        ax.set_facecolor("#FAFAFA")
        total = len(day["series"])

        pcts = [
            100 * sum(1 for o in day["series"] if o["category"] == cat) / total if total > 0 else 0
            for cat in SUMMARY_CATS
        ]

        bars = ax.barh(
            SUMMARY_CATS,
            pcts,
            color=[COLORS.get(c, "#BDBDBD") for c in SUMMARY_CATS],
            edgecolor="white",
            height=0.65,
            alpha=0.9,
        )

        for bar, pct in zip(bars, pcts):
            if pct > 0.5:
                ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height() / 2,
                        f"{pct:.1f} %", va="center", ha="left", fontsize=9, color="#333333")

        label = day["date"].strftime("%d.%m.%Y") if day["date"] else "?"
        ax.set_title(label, fontsize=11, fontweight="bold")
        ax.set_xlabel("Osuus (%)", fontsize=9)
        ax.set_xlim(0, max(max(pcts) * 1.25 + 2, 10))

        for spine in ["top", "right"]:
            ax.spines[spine].set_visible(False)
        ax.spines["left"].set_color("#CCCCCC")
        ax.spines["bottom"].set_color("#CCCCCC")
        ax.grid(axis="x", color="#E0E0E0", linestyle="-", linewidth=0.5)

        ax.text(0.02, 0.98, f"n = {total} havaintoa / ≈ {total} min",
                transform=ax.transAxes, ha="left", va="top", fontsize=8, color="#777777")

    plt.tight_layout()
    return fig


def make_chart3(day_info: list, batch_catalog: dict, person_name: str = ""):
    """
    Työneräkaavio:
    - Tekemisaika ja valmiusaika yläpuolella
    - Apuaika alapuolella työnerän värillä
    - Häiriöaika, muu, taukoaika ja tuntematon alapuolella kiinteillä väreillä
    - Legendassa näkyvät työnerän numero + nimi, jos nimi löytyy Eräluettelo-välilehdeltä
    """
    def build_code_colors(day_info_local):
        unique_codes = []
        seen = set()

        for day in day_info_local:
            for obs in day["series"]:
                if obs["code_key"] not in batch_catalog:
                    continue
                if obs["category"] not in ("Tekemisaika", "Apuaika", "Valmiusaika"):
                    continue
                code_key = obs["code_key"]
                if code_key not in seen:
                    seen.add(code_key)
                    unique_codes.append(code_key)

        unique_codes = sorted(unique_codes, key=sort_code_key)
        cmap = plt.get_cmap("tab20")
        color_map = {}
        for i, code_key in enumerate(unique_codes):
            color_map[code_key] = cmap(i % 20)
        return color_map, unique_codes

    code_colors, unique_codes = build_code_colors(day_info)

    fig, ax = plt.subplots(figsize=(22, 7.2))
    fig.patch.set_facecolor("#F5F5F5")
    ax.set_facecolor("#FAFAFA")

    used_codes = set()

    for day in day_info:
        runs = get_code_runs(day["series"])

        for run in runs:
            x0 = run["x0"]
            width = run["x1"] - run["x0"]
            cat = run["category"]
            code_key = run["code_key"]

            if cat in ("Tekemisaika", "Valmiusaika") and code_key in batch_catalog:
                color = code_colors.get(code_key, COLORS["Tekemisaika"])
                used_codes.add(code_key)
                ax.broken_barh([(x0, width)], (0, 1), facecolor=color, edgecolor="white",
                               linewidth=0.6, alpha=0.95, zorder=3)
            elif cat == "Apuaika" and code_key in batch_catalog:
                color = code_colors.get(code_key, COLORS["Apuaika"])
                used_codes.add(code_key)
                ax.broken_barh([(x0, width)], (-1, 1), facecolor=color, edgecolor="white",
                               linewidth=0.6, alpha=0.95, zorder=3)
            elif cat in ("Häiriöaika", "Muu", "Taukoaika", "Tuntematon") or code_key not in batch_catalog:
                ax.broken_barh([(x0, width)], (-1, 1), facecolor=COLORS.get(cat, COLORS["Tuntematon"]),
                               edgecolor="white", linewidth=0.6, alpha=0.95, zorder=3)

    tick_positions, tick_labels, end_tick_positions = get_xticks_for_day_info(day_info)
    ax.set_xticks(tick_positions)
    ax.set_xticklabels(tick_labels, rotation=45, ha="right", fontsize=8)

    for label, pos in zip(ax.get_xticklabels(), tick_positions):
        if pos in end_tick_positions:
            label.set_color("#C62828")
            label.set_fontweight("bold")
        else:
            label.set_color("#333333")
            label.set_fontweight("normal")

    for day in day_info:
        end_x = day["series"][-1]["x"]
        ax.axvline(end_x, color="#C62828", linestyle=":", linewidth=1.2, zorder=4)

    for day in day_info:
        date_label = day["date"].strftime("%d.%m.%Y") if day["date"] else "?"
        x_start = day["x_start"]
        ax.axvline(x_start, color="#888888", linestyle="--", linewidth=1.2, zorder=4)
        ax.text(x_start + 1.0, -1.45, date_label, rotation=90, rotation_mode="anchor",
                ha="left", va="bottom", fontsize=9, fontweight="bold", color="#444444", zorder=5)

    if len(day_info) > 1:
        for i in range(len(day_info) - 1):
            gap_x = (day_info[i]["x_end"] + day_info[i + 1]["x_start"]) / 2
            ax.axvline(gap_x, color="#BBBBBB", linestyle="--", linewidth=1.0, zorder=2)

    ax.axhline(0, color="#333333", linewidth=1.2, zorder=3)
    ax.set_ylim(-1.6, 2.0)
    ax.set_yticks([-0.5, 0.5])
    ax.set_yticklabels(["Apuaika, häiriöaika ja muu aika", "Tekemisaika"], fontsize=9)
    ax.tick_params(axis="y", length=0)
    ax.set_xlabel("Kellonaika", fontsize=9)

    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.grid(axis="x", color="#E0E0E0", linestyle="-", linewidth=0.5, zorder=0)

    title = "Työneräkaavio – työnerät väreillä, muut ajat alapuolella"
    if person_name:
        title += f" ({person_name})"
    ax.set_title(title, fontsize=12, fontweight="bold", pad=20)

    fig.subplots_adjust(left=0.07, right=0.99, top=0.84, bottom=0.24)
    return fig


# ── Käyttöliittymät ────────────────────────────────────────────────────────


def sanitize_filename(name: str) -> str:
    """Muuttaa tekstin turvalliseksi tiedostonimeksi."""
    cleaned = re.sub(r'[^A-Za-z0-9ÅÄÖåäö._-]+', '_', name.strip())
    return cleaned.strip('._') or "henkilo"


def figure_to_png_bytes(fig) -> bytes:
    """Tallentaa matplotlib-kuvaajan PNG-muotoon tavubufferiin."""
    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    buffer.seek(0)
    return buffer.getvalue()


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Tilastot") -> bytes:
    """Tallentaa DataFramen Excel-muotoon tavubufferiin."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for idx, col in enumerate(df.columns, start=1):
            max_len = max(
                [len(str(col))] + [len(str(v)) for v in df[col].fillna("").tolist()]
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = min(max_len + 2, 40)
    buffer.seek(0)
    return buffer.getvalue()


def build_person_zip(person_name: str, fig1, fig3, fig2, stats_df: pd.DataFrame) -> bytes:
    """Rakentaa ZIP-paketin, jossa ovat kaikki henkilön kuvaajat ja taulukko."""
    safe_name = sanitize_filename(person_name)
    buffer = BytesIO()

    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as zf:
        zf.writestr(f"{safe_name}_kuvaaja1_aikajanakaavio.png", figure_to_png_bytes(fig1))
        zf.writestr(f"{safe_name}_kuvaaja1b_tyonerakaavio.png", figure_to_png_bytes(fig3))
        zf.writestr(f"{safe_name}_kuvaaja2_yhteenveto.png", figure_to_png_bytes(fig2))

        if not stats_df.empty:
            zf.writestr(
                f"{safe_name}_tyoneratilastot.xlsx",
                dataframe_to_excel_bytes(stats_df, sheet_name="Työnerätilastot")
            )

    buffer.seek(0)
    return buffer.getvalue()


def render_combined_code_confirmation(candidates: list):
    """
    Näyttää yhdistelmäkoodien vahvistusosion ja palauttaa
    (approved: bool, overrides: dict).
    """
    if not candidates:
        st.success("Yhdistelmäkoodeja ei löytynyt. Kuvaajat voidaan muodostaa suoraan.")
        return True, {}

    st.subheader("Vahvista yhdistelmäkoodit ennen laskentaa")
    st.write("Ohjelma tunnisti koodeja, joita ei löytynyt suoraan Eräluettelosta. Tarkista ehdotukset ennen kuvaajien muodostamista.")

    category_options = [
        "Tekemisaika", "Apuaika", "Valmiusaika",
        "Häiriöaika", "Muu", "Taukoaika", "Tuntematon"
    ]

    overrides = {}
    with st.form("combined_code_confirmation_form"):
        for cand in candidates:
            code_text = cand["code_text"]
            st.markdown(f"### Koodi {code_text}")
            c1, c2 = st.columns([1, 1])

            with c1:
                st.text_input(
                    "Ohjelman arvio",
                    value=cand["suggested_name"] or "",
                    disabled=True,
                    key=f"suggested_name_{code_text}",
                )
                st.text_input(
                    "Arvion perusta",
                    value=cand["source_label"],
                    disabled=True,
                    key=f"suggested_source_{code_text}",
                )

            with c2:
                st.text_input(
                    "Esimerkkihuomio",
                    value=cand["example_note"] or "",
                    disabled=True,
                    key=f"example_note_{code_text}",
                )
                st.text_input(
                    "Huomioita yhteensä",
                    value=str(cand["note_count"]),
                    disabled=True,
                    key=f"note_count_{code_text}",
                )

            default_category = cand["suggested_category"] if cand["suggested_category"] in category_options else "Tuntematon"
            selected_category = st.selectbox(
                "Valitse aikalaji",
                options=category_options,
                index=category_options.index(default_category),
                key=f"category_override_{code_text}",
            )
            confirmed_name = st.text_input(
                "Vahvistettu nimi",
                value=cand["suggested_name"] or cand["example_note"] or f"Yhdistelmäkoodi {code_text}",
                key=f"name_override_{code_text}",
            )

            overrides[cand["code_key"]] = {
                "name": confirmed_name.strip(),
                "category": selected_category,
                "synthetic": True,
                "source": "user_confirmed",
            }
            st.divider()

        submitted = st.form_submit_button("Hyväksy yhdistelmäkoodit ja muodosta kuvaajat", use_container_width=True)

    return submitted, overrides

def render_person_sections(person_day_infos: dict, batch_catalog: dict, ui_mode: str = "streamlit"):
    if not person_day_infos:
        if ui_mode == "streamlit":
            st.warning("Valituissa tiedostoissa ei ole havaintodataa.")
            st.stop()
        else:
            print("Valituissa tiedostoissa ei ole havaintodataa.")
            sys.exit(0)

    for person_name in sorted(person_day_infos.keys()):
        day_info = person_day_infos[person_name]
        if not day_info:
            continue

        stats_df = compute_batch_run_statistics(day_info, batch_catalog)
        fig1 = make_chart1(day_info, person_name)
        fig3 = make_chart3(day_info, batch_catalog, person_name)
        fig2 = make_chart2(day_info, person_name)

        if ui_mode == "streamlit":
            safe_name = sanitize_filename(person_name)

            st.header(f"Henkilö: {person_name}")
            st.subheader("Kuvaaja 1 – Aikajanakaavio")
            st.pyplot(fig1)

            st.subheader("Kuvaaja 1b – Työneräkaavio")
            st.pyplot(fig3)

            with st.expander("Näytä työneräkaavion selitteet", expanded=False):
                legend_fig = make_chart3_legend_figure(day_info, batch_catalog)
                st.pyplot(legend_fig)
                plt.close(legend_fig)

            st.markdown("**Työneräkohtaiset yhtäjaksoisen tekemisen tilastot**")
            if stats_df.empty:
                st.info("Eräluettelo-välilehdeltä tai yhdistelmäkoodien päätellyistä nimistä ei löytynyt tekemisaikaan kuuluvia työneriä, joita olisi havaittu tässä datassa.")
            else:
                st.dataframe(stats_df, use_container_width=True, hide_index=True)

            st.subheader("Kuvaaja 2 – Yhteenveto päivittäin")
            st.pyplot(fig2)

            st.markdown("**Lataa tiedostot**")
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.download_button(
                    "Kuvaaja 1 (PNG)",
                    data=figure_to_png_bytes(fig1),
                    file_name=f"{safe_name}_kuvaaja1_aikajanakaavio.png",
                    mime="image/png",
                    use_container_width=True,
                    key=f"dl_fig1_{safe_name}",
                )
            with c2:
                st.download_button(
                    "Kuvaaja 1b (PNG)",
                    data=figure_to_png_bytes(fig3),
                    file_name=f"{safe_name}_kuvaaja1b_tyonerakaavio.png",
                    mime="image/png",
                    use_container_width=True,
                    key=f"dl_fig3_{safe_name}",
                )
            with c3:
                st.download_button(
                    "Kuvaaja 2 (PNG)",
                    data=figure_to_png_bytes(fig2),
                    file_name=f"{safe_name}_kuvaaja2_yhteenveto.png",
                    mime="image/png",
                    use_container_width=True,
                    key=f"dl_fig2_{safe_name}",
                )
            with c4:
                if stats_df.empty:
                    st.caption("Ei ladattavaa Excel-taulukkoa")
                else:
                    st.download_button(
                        "Tilastot (Excel)",
                        data=dataframe_to_excel_bytes(stats_df, sheet_name="Työnerätilastot"),
                        file_name=f"{safe_name}_tyoneratilastot.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"dl_xlsx_{safe_name}",
                    )

            st.download_button(
                "Lataa kaikki tämän henkilön tiedostot ZIP-pakettina",
                data=build_person_zip(person_name, fig1, fig3, fig2, stats_df),
                file_name=f"{safe_name}_kuvaajat_ja_tilastot.zip",
                mime="application/zip",
                use_container_width=True,
                key=f"dl_zip_{safe_name}",
            )

            plt.close(fig1)
            plt.close(fig3)
            plt.close(fig2)
        else:
            if not stats_df.empty:
                print(f"\n{person_name} – Työneräkohtaiset yhtäjaksoisen tekemisen tilastot")
                print(stats_df.to_string(index=False))


def run_streamlit():
    st.set_page_config(page_title="Työajan havainnointi", layout="wide")
    st.title("Työajan havainnointi – Kuvaajageneraattori")
    st.write("Lataa yksi tai useampi havainnointi-Excel-tiedosto (.xlsx tai .xlsm).")
    st.write("Jos tiedostossa on useita henkilöitä, jokaiselle henkilösarakkeelle muodostetaan omat kuvaajat.")
    st.write("Tulkinta on kiinteä: C = Henkilö 1, D = Henkilö 2, E = Henkilö 3 ja F = Henkilö 4 kaikissa tiedostoissa.")

    uploaded = st.file_uploader(
        "Valitse Excel-tiedosto(t)",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True,
    )

    if not uploaded:
        st.info("Valitse vähintään yksi tiedosto yllä.")
        st.stop()

    file_datasets = []
    for f in uploaded:
        try:
            file_datasets.append(read_file(f))
        except Exception as e:
            st.error(f"Virhe tiedostossa {f.name}: {e}")

    if not file_datasets:
        st.stop()

    _, candidates, _ = collect_combined_code_candidates(file_datasets)
    approved, overrides = render_combined_code_confirmation(candidates)

    if not approved:
        st.info("Vahvista yhdistelmäkoodit yllä, niin kuvaajat ja tilastot muodostetaan valinnoillasi.")
        st.stop()

    person_day_infos, batch_catalog = build_person_day_infos(file_datasets, catalog_overrides=overrides)
    render_person_sections(person_day_infos, batch_catalog, ui_mode="streamlit")


def run_local():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()
    paths = filedialog.askopenfilenames(
        title="Valitse havainnointi-Excel-tiedosto(t)",
        filetypes=[
            ("Excel-tiedostot", "*.xlsx *.xlsm"),
            ("Kaikki tiedostot", "*.*"),
        ],
    )
    root.destroy()

    if not paths:
        print("Ei valittuja tiedostoja – ohjelma lopetetaan.")
        sys.exit(0)

    file_datasets = []
    errors = []
    for p in paths:
        try:
            file_datasets.append(read_file(p))
        except Exception as e:
            errors.append(f"{p}: {e}")

    if errors:
        root2 = tk.Tk()
        root2.withdraw()
        messagebox.showerror("Virhe tiedostoja luettaessa", "\n".join(errors))
        root2.destroy()

    if not file_datasets:
        sys.exit(1)

    person_day_infos, batch_catalog = build_person_day_infos(file_datasets, catalog_overrides=None)
    render_person_sections(person_day_infos, batch_catalog, ui_mode="local")
    plt.show()


if __name__ == "__main__":
    if STREAMLIT:
        run_streamlit()
    else:
        run_local()
